import re
import requests
import json
from  pprint import pprint
import openpyxl
from models import Address
import sqlite3
from address_classifier_api import call_api


REGION_ID = 8 # Запорізька
DISTRICT_ID = 950 # Мелітопольський
CITY_ID = 4633 # Мелітополь
# CITY_ID = 4454 # Костянтинівка
 
def xls_file_read(name_xls):
    wb_obj = openpyxl.load_workbook(name_xls)
    sheet = wb_obj.active
    table = []
    
    for row in sheet.iter_rows(1, sheet.max_row):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        # print(cell.value)
        table.append(row_data)
    
    return table

def main():
    query = {
        'region_id': REGION_ID, 
        'district_id': DISTRICT_ID,
        # 'city_id': CITY_ID
    }
    parameters = {'region_id': REGION_ID, 'district_id': DISTRICT_ID, 'city_id': CITY_ID}
    res_json = call_api('get_street', parameters=parameters)
    if res_json:
        write_result_2_json('streets_'+str(CITY_ID)+'.json', res_json)

def read_citys():
    with open('data.json', 'r', encoding='utf-8') as f:
        citys_data =  json.load(f)
        # json.dump(city_list, f, ensure_ascii=False, indent=4)
    citys = {}
    for city in citys_data:
        citys[city['CITY_UA']] = city['CITY_ID']
    return citys

def read_streets():
    with open('streets.json', 'r', encoding='utf-8') as f:
        streets_data =  json.load(f)
    streets = streets_data['Entries']['Entry']
    # for street in streets_data:
    #     streets[street['streettype_ua']+'_'+street['street_ua'].lower()] = street['street_id']
    return streets

def get_short_streets(streets):
    short_streets = []
    st_list = []
    for i, st in enumerate(streets):
        # import ipdb; ipdb.set_trace()
        st_list.append((st['STREETTYPE_UA']+' '+st['STREET_UA'], i))
    short_streets = dict(st_list)
    return short_streets

def parse_adress(table):
    adress_list = []
    for ad in table:
        # костыль для Олиной работы 
        res_street = ad[0].lower().replace('рєпина', 'рєпіна')
        # конец костыля
        adress_list.append(Address(res_street))
    return adress_list

def find_street_id(ad, streets:dict, short_streets:dict) -> str:
    def search_in_part(find_text, short_part:dict):
        part_result=[]
        regex = re.compile(
            r'\b'+find_text,
            flags=re.IGNORECASE | re.VERBOSE
        )
        for k, v in short_part.items():
            if regex.search(k):
                part_result.append((k, v))
        return part_result

    name_street = ad.street+' '+ad.streettype
    mass_name_part = name_street.split(' ')
    results = []
    if len(mass_name_part)>1:
        buffer_result = []
        search_part = short_streets
        for i, part in enumerate(mass_name_part):
            buffer_result = search_in_part(part, search_part)
            search_part = dict(buffer_result)
        results = buffer_result
    new_res = []
    for elem in results:
        new_res.append({'street': elem[0], 'index': elem[1], 'street_id': streets[elem[1]]['STREET_ID']})
    return new_res

def write_result_2_json(name_file, json_data):
    with open(name_file, 'w', encoding='utf-8') as f:
        json.dump(json_data, f, ensure_ascii=False, indent=4)


def find_post_code(json_result):
    # [{'HOUSENUMBER_UA': '26', 'POSTCODE': '72312', 'STREET_ID': '405156'}]
    street_postcodes = {}
    entries = json_result['Entries'].get('Entry')
    for entry in entries:
        # import ipdb; ipdb.set_trace()
        val = street_postcodes.get(entry['POSTCODE'], '/')
        val += entry['HOUSENUMBER_UA']+'/'
        street_postcodes[entry['POSTCODE']]=val
    return street_postcodes

def list_to_str(list_result:list) -> str:
    res = {}
    for value in list_result:
        if not isinstance(value, dict):
        # import ipdb; ipdb.set_trace()
            val = res.get(value, '0')
            res[value]=val
        else:
            str_val = ''
            for k,v in value.items():
                str_val += k+' :'+v+'\n'
            return str_val
    str_val = repr(res.keys)
    str_val = ''.join(['%s' % (key) for (key, value) in res.items()])
    # import ipdb; ipdb.set_trace()
    return str_val

if __name__=="__main__":
    main()
    exit()
    # citys_list = read_citys()

    streets = read_streets()
    short_streets = get_short_streets(streets)

    # table = xls_file_read('./input/adress.xlsx')
    table = xls_file_read('./input/adress_оля_вторник_short.xlsx')
    
    # import ipdb; ipdb.set_trace()
    
    # exit()
    print('parse_adress')

    adress_list = parse_adress(table)
    adress_list_postcode = []
    
    
    print('find_street_id')
    
    for i, ad in enumerate(adress_list):
        adress_list_postcode.append({'raw_street': table[i], 'obj_adress': ad, 'dict_street_id': find_street_id(ad, streets, short_streets)})
    
    print('find_postcode:')
    len_adress = len(adress_list_postcode)

    for i, ad in enumerate(adress_list_postcode):
        if i%10==0:
            print(f'{len_adress}/{i}')
        find_postcodes = []
        dont_found_postcodes = []
        for street_ids in ad['dict_street_id']:
            street_id = street_ids['street_id']
            building = ad['obj_adress'].building
            # print('raw_street')
            # pprint(ad['raw_street'])
            # print(f'street_id: {street_id} , build: {building}')
            # print('obj_adress')
            # pprint(ad['obj_adress'])
            # print('json_result')
            json_result = call_api('get_house_number', street_id, building)
            #
            entry = json_result['Entries'].get('Entry')
            if entry:
                res = entry[0]['POSTCODE']
                find_postcodes.append(res)
            else:
                json_result = call_api('get_house_number', street_id, '')
                res = find_post_code(json_result)
                dont_found_postcodes.append(res)
            
            # pprint(find_postcodes)
            # print('='*20)
            # find_postcodes.append(call_api('get_house_number', street_id, building)['Entries']['Entry']['POSTCODE'])
            # pprint(call_api('get_house_number', street_id, building)['Entries']['Entry'])
        ad['find_postcodes']=find_postcodes
        ad['dont_found_postcodes']=dont_found_postcodes
    # pprint(adress_list_postcode)
    
    print('write to excel')
    excel_file = openpyxl.Workbook()
    excel_sheet = excel_file.create_sheet(title='result postcode', index=0)
    
    for row in adress_list_postcode:
        excel_sheet.append([row['raw_street'][0], list_to_str(row['find_postcodes']), list_to_str(row['dont_found_postcodes'])])
    # main()

    file_name = 'adress_оля_вторник_short'
    excel_file.save(f'./output/result {file_name}.xlsx')
    print('done')

    
